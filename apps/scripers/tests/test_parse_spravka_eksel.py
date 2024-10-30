#import xlrd
#pip install xlrd==1.2.0
import openpyxl
from datetime import datetime 
from pytz import timezone, utc

import calendar
from decimal import Decimal
if __name__ == "__main__":
    # Open Excel file
    workbook = openpyxl.load_workbook('./Справка м.07.2024_ГОФРИЛО КО ЕООД.xlsx')
    #print(workbook.get_sheet_names())
    print(workbook)
    localtimezone = timezone("Europe/Sofia")

    for sheet_name in workbook.sheetnames:
        print(sheet_name)
        sheet = workbook[sheet_name]  # Access sheet by name
        print(sheet)
        
        month        = sheet["B1"].value
        owner_legal  = sheet["C1"].value
        factory_name = sheet["C2"].value
        some_id      = sheet["C3"].value
        num_days     = calendar.monthrange(month.year, month.month)[1]

        
        print(month)
        print(type(month))
        print(owner_legal)
        print(factory_name)
        print(some_id)
        
        production_label = sheet["B4"]
        print(production_label)
        
        print("Num Days %d" % num_days)
        for h in range(24):
            v = sheet.cell(row = 4, column = 3 + h).value
            
            print(v)
            
        for d in range(num_days):
            date_label = sheet.cell(row = 5 + d, column = 2).value
            #print(date_label)
            for h in range(24):
                production = sheet.cell(row = 5 + d, column = 3 + h).value
                price = sheet.cell(row = 5 + d, column = 3 + h + 27).value
                d1 = datetime(year=month.year, month=month.month, day=d+1, hour=h)
                d1 = localtimezone.localize(d1)
                d1 = d1.astimezone(utc)                
                #print(type(production))
                p = Decimal("%0.4f" % production) * Decimal(1000)
                price = Decimal("%0.2f" % price)
                print(str(d1) + ' -> ' + str(p) + ' kWh @ ' + str(price) )
                
    
