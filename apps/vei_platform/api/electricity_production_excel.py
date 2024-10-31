# import xlrd
# pip install xlrd==1.2.0
import openpyxl
from datetime import datetime, timedelta
from pytz import timezone, utc

import calendar
from decimal import Decimal
from vei_platform.models.factory import ElectricityFactory
#from vei_platform.models.legal import LegalEntity

def process_excel_report(filename="./Справка м.07.2024_ГОФРИЛО КО ЕООД.xlsx" ):
    # Open Excel file
    workbook = openpyxl.load_workbook(filename)
    factories = []

    for sheet_name in workbook.sheetnames:
        # print(sheet_name)
        sheet = workbook[sheet_name]  # Access sheet by name
        # print(sheet)

        month = sheet["B1"].value
        owner_legal = sheet["C1"].value
        factory_name = sheet["C2"].value
        some_id = sheet["C3"].value
        num_days = calendar.monthrange(month.year, month.month)[1]
        
        factory_object = ElectricityFactory.objects.filter(name=factory_name).first()
        if factory_object:
            factory_slug = factory_object.slug    
            timezone_label = factory_object.get_requested_timezone()
        else:
            factory_slug = None
            timezone_label = "UTC"
        
        localtimezone = timezone(timezone_label)
        

        # print(month)
        # print(type(month))
        # print(owner_legal)
        # print(factory_name)
        # print(some_id)

        production_label = sheet["B4"].value
        # print(production_label)
        if production_label != "Количество\nМВтч":
            print("Expected B4 to be 'Количество\nМВтч'")

        # print("Num Days %d" % num_days)
        # check hour labels
        for h in range(24):
            v = sheet.cell(row=4, column=3 + h).value
            if v != h + 1:
                print("ERROROROROR")

        production_in_kwh = []
        prices = []

        for d in range(num_days):
            date_label = sheet.cell(row=5 + d, column=2).value
            # print(date_label)
            for h in range(24):
                production = sheet.cell(row=5 + d, column=3 + h).value
                price = sheet.cell(row=5 + d, column=3 + h + 27).value
                d1 = datetime(year=month.year, month=month.month, day=d + 1, hour=h)
                d1 = localtimezone.localize(d1)
                d1 = d1.astimezone(utc)
                start_interval = d1.strftime("%Y-%m-%dT%H:%M:%S%z")
                end_interval = (d1 + timedelta(hours=1)).strftime("%Y-%m-%dT%H:%M:%S%z")
                # print(type(production))
                energy_in_kwh = Decimal("%0.4f" % production) * Decimal(1000)
                price = Decimal("%0.2f" % price)
                # print(str(d1) + ' -> ' + str(p) + ' kWh @ ' + str(price) )
                production_in_kwh.append(
                    {
                        "factory": factory_slug,
                        "energy_in_kwh": energy_in_kwh,
                        "start_interval": start_interval,
                        "end_interval": end_interval,
                    }
                )

                prices.append(
                    {
                        "factory": factory_slug,
                        "price": price,
                        "start_interval": start_interval,
                        "end_interval": end_interval,
                    }
                )

        res = {
            "factory_name": factory_name,
            "factory_slug": factory_slug,
            "timezone": timezone_label,
            "factory_id": some_id,
            "legal_entity": owner_legal,
            "month": month.month,
            "year": month.year,
            "production_in_kwh": production_in_kwh,
            "prices": prices,
            "currency": "BGN",
        }
        factories.append(res)

    return factories
