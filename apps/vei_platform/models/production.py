from django.db import models
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from .factory import ElectricityFactory
from .restricted_file_field import RestrictedFileField
from django.contrib.auth.models import User
from djmoney.models.fields import Decimal, MoneyField
from djmoney.money import Money
from djmoney.contrib.exchange.models import convert_money

from .electricity_price import ElectricityPricePlan, ElectricityPrice


class ElectricityFactoryProduction(models.Model):
    factory = models.ForeignKey(
        ElectricityFactory, null=True, blank=True, on_delete=models.CASCADE
    )

    start_interval = models.DateTimeField(
        blank=False,
        null=False,
        db_index=True,
        default=datetime(
            year=2024, month=1, day=1, hour=0, minute=0, tzinfo=timezone.utc
        ),
    )

    end_interval = models.DateTimeField(
        blank=False,
        null=False,
        db_index=True,
        default=datetime(
            year=2024, month=1, day=1, hour=1, minute=0, tzinfo=timezone.utc
        ),
    )

    energy_in_kwh = models.DecimalField(
        max_digits=14, decimal_places=2, default=Decimal(0)
    )

    reported_price_per_mwh = MoneyField(
        max_digits=14,
        decimal_places=2,
        default_currency="BGN",
        default=None,
        blank=True,
        null=True,
    )

    def __str__(self) -> str:
        return "%s %s kWh" % (
            self.start_interval.strftime("%y-%m-%d %H:%M %Z"),
            self.energy_in_kwh,
        )


def user_document_upload_directory_path(instance, filename):
    # file will be uploaded to MEDIA_ROOT/user_<id>/<filename>
    res = "documents/user_{0}/{1}".format(str(instance.owner), filename)
    print("-------->>>>>>>" + res)
    return res


class ElectricityFactoryUploadedProductionReport(models.Model):
    docfile = RestrictedFileField(
        upload_to=user_document_upload_directory_path,
        content_types=[
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ],
        max_upload_size=6 * 1024 * 1024,  # 5 MB
        default=None,
        null=False,
        blank=False,
    )
    owner = models.ForeignKey(User, null=False, blank=False, on_delete=models.CASCADE)
    created_at = models.DateTimeField(auto_now_add=True)
    updated_at = models.DateTimeField(auto_now=True)


class ElectricityFactoryProductionReport(models.Model):
    months_selection = (
        (1, "January"),
        (2, "February"),
        (3, "March"),
        (4, "April"),
        (5, "May"),
        (6, "June"),
        (7, "July"),
        (8, "August"),
        (9, "September"),
        (10, "October"),
        (11, "November"),
        (12, "December"),
    )

    report = models.ForeignKey(
        ElectricityFactoryUploadedProductionReport,
        null=False,
        blank=False,
        on_delete=models.CASCADE,
    )
    factory = models.ForeignKey(
        ElectricityFactory, blank=False, null=False, on_delete=models.CASCADE
    )
    year = models.IntegerField("Year")
    month = models.IntegerField("Month", choices=months_selection)

    class Meta:
        unique_together = [["factory", "year", "month"]]


# import xlrd
# pip install xlrd==1.2.0
import openpyxl
from datetime import datetime, timedelta
from pytz import timezone, utc

import calendar
from decimal import Decimal
from vei_platform.models.factory import ElectricityFactory
from django.utils.translation import gettext as _
import re


def exctract_timezone_currency_unit(sheet):
    if "IBEX_DAM_Price" == sheet["B1"].value and "лв./MWh" == sheet["B2"].value:
        r = re.match(r"Date\/(\S+)", sheet["A1"].value)
        if r:
            timezone_label = r.group(1)
            # print(timezone_label)
            num_factories = count_factories_old_format(sheet)
            # print(num_factories)
            # print("A1 = ", sheet["A1"].value)
            # print("A2 = ", sheet["A2"].value)
            if num_factories > 0:
                return timezone_label, "BGN", "MWh"
            # return num_factories > 0
    return None, None, None


def count_factories_old_format(sheet):
    res = 0
    for h in range(1024):
        v1 = sheet.cell(row=2, column=4 + 2 * h + 0).value
        v2 = sheet.cell(row=2, column=4 + 2 * h + 1).value
        if v1 == "Производство, MWh" and v2 == "Сума, лв.":
            res = res + 1
        else:
            # print ("v1 = %s", v1)
            # print ("v2 = %s", v2)
            break
    return res


def find_factory(search_id):
    for id in search_id.split(" "):
        try:
            return ElectricityFactory.objects.get(factory_code=id)
        except ElectricityFactory.DoesNotExist:
            pass
    try:
        return ElectricityFactory.objects.get(name=search_id)
    except ElectricityFactory.DoesNotExist:
        pass
    return None


def add_production(
    production_in_kwh,
    factory_slug,
    production_in_mwh,
    price,
    start_interval: datetime,
    end_interval=None,
):
    start = start_interval
    end = start + timedelta(hours=1) if end_interval is None else end_interval

    energy_in_mwh = (
        production_in_mwh
        if isinstance(production_in_mwh, Decimal)
        else Decimal("%0.6f" % production_in_mwh)
    )
    energy_in_kwh = energy_in_mwh * Decimal(1000)
    start_str = start.strftime("%Y-%m-%dT%H:%M:%S%z")
    end_str = end.strftime("%Y-%m-%dT%H:%M:%S%z")

    price_decimal = price if isinstance(price, Decimal) else Decimal("%0.2f" % price)

    production_in_kwh.append(
        {
            "factory": factory_slug,
            "energy_in_kwh": energy_in_kwh,
            "reported_price_per_mwh": price_decimal,
            "reported_price_per_mwh_currency": "BGN",
            "start_interval": start_str,
            "end_interval": end_str,
        }
    )


from djmoney.contrib.exchange.exceptions import MissingRate


def convert_to_correct_currency(price, currency, plan_currency=None):
    if plan_currency is None:
        return price
    # price = Decimal(str(sheet.cell(row=3 + r, column=2).value))
    try:
        p0 = Money(price, currency)
        p1 = convert_money(p0, plan_currency)
        res = p1.amount.quantize(Decimal(".01"), rounding=ROUND_HALF_UP)
        # print(p0, " -> ", p1, ' ', res)
        return res
    except MissingRate:
        # print(p0, ' ->', p1)
        if "BGN" == currency and plan_currency == "EUR":
            price = price * Decimal("0.51129188")
            price = price.quantize(Decimal(".01"), rounding=ROUND_HALF_UP)
            return price


def extract_factory_production(sheet, h, timezone_label, currency, unit):
    errors = []
    factory_name = None
    factory_slug = None
    factory_id = None
    owner_legal = None
    report_period_start = None
    production_in_kwh = []
    prices = []
    localtimezone = timezone(timezone_label)

    v1 = sheet.cell(row=2, column=4 + 2 * h + 0).value
    v2 = sheet.cell(row=2, column=4 + 2 * h + 1).value
    start_interval, end_interval = None, None
    if v1 == "Производство, MWh" and v2 == "Сума, лв.":

        search_text = sheet.cell(row=1, column=4 + 2 * h + 0).value
        f = find_factory(search_text)
        factory_slug = f.slug if f is not None else None
        factory_name = f.name if f is not None else None
        factory_id = f.factory_code if f is not None else search_text.split(" ")[0]
        plan_slug = f.plan.slug if f is not None else None
        plan_currency = None
        if plan_slug:
            try:
                plan = ElectricityPricePlan.objects.get(slug=plan_slug)
                plan_currency = plan.currency
            except ElectricityPricePlan.DoesNotExist:
                pass

        for r in range(4 * 24 * 31 + 10):
            start = sheet.cell(row=3 + r, column=1).value
            # end interval check
            # if end_interval is not None and end_interval == start:
            if end_interval is not None and num_intervals == 4:
                add_production(
                    production_in_kwh,
                    factory_slug,
                    total_prod,
                    price,
                    start_interval=start_interval,
                )
                start_interval = end_interval
                end_interval = start_interval + timedelta(hours=1)
                total_prod = Decimal(0)
                num_intervals = 0

            if not start:
                break

            # start inteval check
            price = Decimal(str(sheet.cell(row=3 + r, column=2).value))
            price = convert_to_correct_currency(price, currency, plan_currency)
            prod = Decimal(str(sheet.cell(row=3 + r, column=4 + 2 * h + 0).value))

            if start_interval is None:
                start_interval = start
                report_period_start = start
                report_period_start = localtimezone.localize(report_period_start)
                start_interval = localtimezone.localize(start_interval)
                start_interval = start_interval.astimezone(utc)
                end_interval = start_interval + timedelta(hours=1)
                total_prod = Decimal(0)
                num_intervals = 0

            if start_interval is not None:
                total_prod += prod
                num_intervals += 1

    res = {
        "factory_name": factory_name,
        "factory_slug": factory_slug,
        "timezone": timezone_label,
        "factory_id": factory_id,
        "legal_entity": owner_legal,
        "month": report_period_start.month,
        "year": report_period_start.year,
        "production_in_kwh": production_in_kwh,
        "errors": errors,
    }
    return res


def process_old_excel_report(sheet, timezone_label, currency, unit):
    factories = []
    num_factories = count_factories_old_format(sheet)
    for num_factory in range(num_factories):
        factories.append(
            extract_factory_production(
                sheet, num_factory, timezone_label, currency, unit
            )
        )
    return factories


def num_hours_in_day(year, month, day, localtimezone):
    d0 = datetime(year=year, month=month, day=day, hour=0)
    d0 = localtimezone.localize(d0)
    d0 = d0.astimezone(utc)
    num_days = calendar.monthrange(year, month)[1]
    if day + 1 <= num_days:
        d1 = datetime(year=year, month=month, day=day + 1, hour=0)
    elif month < 12:
        d1 = datetime(year=year, month=month + 1, day=1, hour=0)
    else:
        d1 = datetime(year=year + 1, month=1, day=1)
    d1 = localtimezone.localize(d1)
    d1 = d1.astimezone(utc)
    return int(d1.timestamp() - d0.timestamp()) // 3600


def process_excel_report(filename):
    # Open Excel file
    workbook = openpyxl.load_workbook(filename)
    factories = []

    for sheet_name in workbook.sheetnames:
        # print(sheet_name)
        sheet = workbook[sheet_name]  # Access sheet by name
        timezone_label, currency, unit = exctract_timezone_currency_unit(sheet)
        if timezone_label and currency and unit:
            return process_old_excel_report(sheet, timezone_label, currency, unit)

        errors = []
        # print(sheet)

        month = sheet["B1"].value
        owner_legal = sheet["C1"].value
        factory_name = sheet["C2"].value
        some_id = sheet["C3"].value

        if isinstance(month, datetime):
            num_days = calendar.monthrange(month.year, month.month)[1]
        else:
            errors.append(
                _("Report does not contain period in B1 cell, found %s") % str(month)
            )
            num_days = 0

        production_label = sheet["B4"].value
        if production_label != "Количество\nМВтч":
            errors.append("Expected B4 to be 'Количество\nМВтч'")
            res = {
                "errors": errors,
            }
            factories.append(res)
            continue

        expected_price_label = "Цена\nЛева/МВтч"
        currency = "BGN"
        max_hours_in_day = 0
        for c in range(12):
            if sheet.cell(row=4, column=2 + 27 + c).value == expected_price_label:
                max_hours_in_day = 24 + c
                break

        # print(max_hours_in_day)

        factory_object = find_factory(factory_name)

        # factory_object = ElectricityFactory.objects.filter(name=factory_name).first()
        if factory_object:
            factory_slug = factory_object.slug
            timezone_label = factory_object.get_requested_timezone()
            plan_slug = factory_object.plan.slug
            plan_currency = factory_object.plan.currency
        else:
            errors.append(_("Factory with name `%s` not found") % factory_name)
            factory_slug = None
            timezone_label = "EET"
            plan_slug = None

        localtimezone = timezone(timezone_label)

        # print("Num Days %d" % num_days)
        # check hour labels
        for h in range(max_hours_in_day):
            v = sheet.cell(row=4, column=3 + h).value
            c = sheet.cell(row=4, column=3 + h)
            if v != h + 1:
                errors.append(
                    "Expected value %d in cell %s%d" % (h + 1, c.column_letter, c.row)
                )

        production_in_kwh = []
        production = ElectricityFactoryProduction.objects.filter(factory=factory_object)
        d0 = datetime(year=month.year, month=month.month, day=1, hour=0)
        d0 = localtimezone.localize(d0)
        d0 = d0.astimezone(utc)
        for d in range(num_days):
            date_label = sheet.cell(row=5 + d, column=2).value
            day = d + 1
            n = num_hours_in_day(month.year, month.month, day, localtimezone)
            for h in range(n):
                production = sheet.cell(row=5 + d, column=3 + h).value
                # print(production)
                price = sheet.cell(
                    row=5 + d, column=max_hours_in_day + 3 + h + 27 - 24
                ).value
                price = convert_to_correct_currency(price, currency, plan_currency)
                add_production(
                    production_in_kwh,
                    factory_slug,
                    production,
                    price,
                    start_interval=d0,
                )
                d0 = d0 + timedelta(hours=1)
                # if n == 25:
                #    print(d0, ' -> %.2f' % price)
        res = {
            "factory_name": factory_name,
            "factory_slug": factory_slug,
            "timezone": timezone_label,
            "factory_id": some_id,
            "legal_entity": owner_legal,
            "month": month.month,
            "year": month.year,
            "production_in_kwh": production_in_kwh,
            "errors": errors,
        }
        factories.append(res)

    return factories
